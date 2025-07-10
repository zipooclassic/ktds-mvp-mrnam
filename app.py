from flask import Flask, render_template, redirect, url_for, request, jsonify, send_file
import sqlite3
import json
from datetime import datetime, timedelta
import code_validator  # Assuming this is the module with your validation logic
from code_generator_service import get_code_generator
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 코드 생성 서비스 인스턴스 가져오기
code_generator_service = get_code_generator()

if not code_generator_service:
    logger.error("CodeGeneratorService 초기화 실패")
else:
    logger.info("CodeGeneratorService 초기화 성공")

def get_db_connection():
    return sqlite3.connect('validation_results.db')

@app.route('/')
def index():
    # 기본 홈페이지는 Code Inspection으로 리다이렉트
    return redirect(url_for('code_inspection'))

@app.route('/code-generation')
def code_generation():
    return render_template('code_generation.html')

@app.route('/code-inspection')
def code_inspection():
    conn = get_db_connection()
    thirty_days_ago_str = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')

    # 기본값 설정
    stats = {
        "total_commits": 0,
        "total_violations": 0,
        "total_resolutions": 0,
        "avg_resolution_days": 0
    }
    
    author_stats = {
        "labels": [],
        "violations": [],
        "resolutions": [],
    }
    
    trend_stats = {
        "labels": [],
        "data": [],
    }
    
    violation_type_stats = {
        "labels": ["명명규칙", "CQRS 패턴", "클래스 구조", "기타"],
        "data": [40, 30, 20, 10]
    }
    
    daily_resolution_stats = {
        "labels": [],
        "data": []
    }
    
    history_log = []
    
    try:
        # --- 1. KPI 통계 계산 ---
        # "코드 가이드라인을 모두 준수했습니다." 메시지는 실제 위반사항이 아니므로 제외
        stats = {
            "total_commits": conn.execute("SELECT COUNT(DISTINCT violation_commit_hash) FROM validations WHERE created_at >= ? AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'", (thirty_days_ago_str,)).fetchone()[0] or 0,
            "total_violations": conn.execute("SELECT COUNT(id) FROM validations WHERE created_at >= ? AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'", (thirty_days_ago_str,)).fetchone()[0] or 0,
            "total_resolutions": conn.execute("SELECT COUNT(id) FROM validations WHERE status = 'closed' AND resolved_at >= ? AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'", (thirty_days_ago_str,)).fetchone()[0] or 0,
        }
        avg_days_row = conn.execute("SELECT AVG(julianday(resolved_at) - julianday(created_at)) FROM validations WHERE status = 'closed' AND resolved_at IS NOT NULL AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'").fetchone()
        stats['avg_resolution_days'] = round(avg_days_row[0], 1) if avg_days_row and avg_days_row[0] is not None else 0.0

        # --- 2. 개발자별 통계 (차트용) ---
        author_violations = dict(conn.execute("SELECT author_name, COUNT(id) FROM validations WHERE violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%' GROUP BY author_name").fetchall())
        author_resolutions = dict(conn.execute("SELECT resolved_by_author, COUNT(id) FROM validations WHERE status = 'closed' AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%' GROUP BY resolved_by_author").fetchall())
        all_authors = sorted(list(set(author_violations.keys()) | set(author_resolutions.keys())))
        
        if all_authors:
            author_stats = {
                "labels": all_authors,
                "violations": [author_violations.get(author, 0) for author in all_authors],
                "resolutions": [author_resolutions.get(author, 0) for author in all_authors],
            }

        # --- 3. 월별 위반 트렌드 (차트용) ---
        trend_data = conn.execute("SELECT strftime('%Y-%m', created_at) as month, COUNT(id) FROM validations WHERE violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%' GROUP BY month ORDER BY month").fetchall()
        if trend_data:
            trend_stats = {
                "labels": [row[0] for row in trend_data],
                "data": [row[1] for row in trend_data],
            }
        
        # --- 4. 위반 유형별 분포 (상세 데이터 포함) ---
        # 키워드 기반으로 위반 유형별 데이터 조회
        violation_type_stats = {
            "labels": ["명명규칙", "CQRS 패턴", "클래스 구조", "기타"],
            "data": [0, 0, 0, 0],
            "details": {
                "명명규칙": [],
                "CQRS 패턴": [],
                "클래스 구조": [],
                "기타": []
            }
        }
        
        # 각 위반 유형별로 직접 데이터베이스 조회
        
        # 1. 명명규칙 위반
        naming_violations = conn.execute("""
            SELECT violation_details, file_path, author_name, created_at, status, violation_commit_hash
            FROM validations 
            WHERE created_at >= ? 
            AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            AND (violation_details LIKE '%명명%' OR violation_details LIKE '%네이밍%' OR violation_details LIKE '%이름%' 
                 OR violation_details LIKE '%Naming%' OR violation_details LIKE '%name%')
            ORDER BY created_at DESC
        """, (thirty_days_ago_str,)).fetchall()
        
        for row in naming_violations:
            violation_type_stats["details"]["명명규칙"].append({
                "file_path": row[1] if row[1] else "",
                "author": row[2] if row[2] else "",
                "created_at": row[3] if row[3] else "",
                "status": row[4] if row[4] else "",
                "commit_hash": row[5] if row[5] else "",
                "description": row[0][:200] + "..." if row[0] and len(row[0]) > 200 else (row[0] or "")
            })
        
        # 2. CQRS 패턴 위반
        cqrs_violations = conn.execute("""
            SELECT violation_details, file_path, author_name, created_at, status, violation_commit_hash
            FROM validations 
            WHERE created_at >= ? 
            AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            AND (violation_details LIKE '%CQRS%' OR violation_details LIKE '%Command%' OR violation_details LIKE '%Query%'
                 OR violation_details LIKE '%Cmd%' OR violation_details LIKE '%Qry%')
            ORDER BY created_at DESC
        """, (thirty_days_ago_str,)).fetchall()
        
        for row in cqrs_violations:
            violation_type_stats["details"]["CQRS 패턴"].append({
                "file_path": row[1] if row[1] else "",
                "author": row[2] if row[2] else "",
                "created_at": row[3] if row[3] else "",
                "status": row[4] if row[4] else "",
                "commit_hash": row[5] if row[5] else "",
                "description": row[0][:200] + "..." if row[0] and len(row[0]) > 200 else (row[0] or "")
            })
        
        # 3. 클래스 구조 위반
        class_violations = conn.execute("""
            SELECT violation_details, file_path, author_name, created_at, status, violation_commit_hash
            FROM validations 
            WHERE created_at >= ? 
            AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            AND (violation_details LIKE '%클래스%' OR violation_details LIKE '%구조%' OR violation_details LIKE '%아키텍처%'
                 OR violation_details LIKE '%class%' OR violation_details LIKE '%structure%' OR violation_details LIKE '%responsibility%')
            ORDER BY created_at DESC
        """, (thirty_days_ago_str,)).fetchall()
        
        for row in class_violations:
            violation_type_stats["details"]["클래스 구조"].append({
                "file_path": row[1] if row[1] else "",
                "author": row[2] if row[2] else "",
                "created_at": row[3] if row[3] else "",
                "status": row[4] if row[4] else "",
                "commit_hash": row[5] if row[5] else "",
                "description": row[0][:200] + "..." if row[0] and len(row[0]) > 200 else (row[0] or "")
            })
        
        # 4. 기타 위반 (위의 3개 카테고리에 속하지 않는 것들)
        other_violations = conn.execute("""
            SELECT violation_details, file_path, author_name, created_at, status, violation_commit_hash
            FROM validations 
            WHERE created_at >= ? 
            AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            AND NOT (violation_details LIKE '%명명%' OR violation_details LIKE '%네이밍%' OR violation_details LIKE '%이름%' 
                     OR violation_details LIKE '%Naming%' OR violation_details LIKE '%name%')
            AND NOT (violation_details LIKE '%CQRS%' OR violation_details LIKE '%Command%' OR violation_details LIKE '%Query%'
                     OR violation_details LIKE '%Cmd%' OR violation_details LIKE '%Qry%')
            AND NOT (violation_details LIKE '%클래스%' OR violation_details LIKE '%구조%' OR violation_details LIKE '%아키텍처%'
                     OR violation_details LIKE '%class%' OR violation_details LIKE '%structure%' OR violation_details LIKE '%responsibility%')
            ORDER BY created_at DESC
        """, (thirty_days_ago_str,)).fetchall()
        
        for row in other_violations:
            violation_type_stats["details"]["기타"].append({
                "file_path": row[1] if row[1] else "",
                "author": row[2] if row[2] else "",
                "created_at": row[3] if row[3] else "",
                "status": row[4] if row[4] else "",
                "commit_hash": row[5] if row[5] else "",
                "description": row[0][:200] + "..." if row[0] and len(row[0]) > 200 else (row[0] or "")
            })
        
        # 실제 데이터로 차트 데이터 업데이트
        violation_type_stats["data"] = [
            len(violation_type_stats["details"]["명명규칙"]),
            len(violation_type_stats["details"]["CQRS 패턴"]),
            len(violation_type_stats["details"]["클래스 구조"]),
            len(violation_type_stats["details"]["기타"])
        ]
        
        # 디버깅용 출력
        print(f"위반 유형별 통계:")
        print(f"  - 명명규칙: {len(violation_type_stats['details']['명명규칙'])}건")
        print(f"  - CQRS 패턴: {len(violation_type_stats['details']['CQRS 패턴'])}건")
        print(f"  - 클래스 구조: {len(violation_type_stats['details']['클래스 구조'])}건")
        print(f"  - 기타: {len(violation_type_stats['details']['기타'])}건")
        
        # --- 5. 일별 조치 현황 (최근 7일) ---
        daily_resolution_data = []
        daily_resolution_labels = []
        for i in range(6, -1, -1):
            date = datetime.now() - timedelta(days=i)
            date_str = date.strftime('%Y-%m-%d')
            count = conn.execute("SELECT COUNT(id) FROM validations WHERE status = 'closed' AND DATE(resolved_at) = ? AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'", (date_str,)).fetchone()[0] or 0
            daily_resolution_data.append(count)
            daily_resolution_labels.append(date.strftime('%m/%d'))
        
        daily_resolution_stats = {
            "labels": daily_resolution_labels,
            "data": daily_resolution_data
        }
        
        # --- 6. 최근 활동 로그 ---
        # "코드 가이드라인을 모두 준수했습니다." 메시지 제외
        recent_violations = conn.execute("SELECT created_at, 'open' as status, file_path, author_name, violation_commit_hash, violation_details FROM validations WHERE violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%' ORDER BY created_at DESC LIMIT 10").fetchall()
        recent_resolutions = conn.execute("SELECT resolved_at, 'closed' as status, file_path, resolved_by_author, resolved_commit_hash, 'Fixed' as details FROM validations WHERE status = 'closed' AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%' ORDER BY resolved_at DESC LIMIT 10").fetchall()
        
        # 두 리스트를 합치고 시간 순으로 정렬
        for row in recent_violations:
            if row[0]:  # created_at이 None이 아닌 경우
                try:
                    timestamp = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
                    history_log.append({
                        'timestamp': timestamp, 
                        'status': row[1], 
                        'file_path': row[2], 
                        'author': row[3], 
                        'commit_hash': row[4], 
                        'details': row[5]
                    })
                except ValueError:
                    continue
                    
        for row in recent_resolutions:
            if row[0]:  # resolved_at이 None이 아닌 경우
                try:
                    timestamp = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
                    history_log.append({
                        'timestamp': timestamp, 
                        'status': row[1], 
                        'file_path': row[2], 
                        'author': row[3], 
                        'commit_hash': row[4], 
                        'details': row[5]
                    })
                except ValueError:
                    continue
        
        history_log.sort(key=lambda x: x['timestamp'], reverse=True)
        
    except Exception as e:
        print(f"데이터베이스 조회 중 오류: {e}")
    finally:
        conn.close()
    
    return render_template('code_inspection.html', 
                         stats=stats, 
                         author_stats=author_stats, 
                         trend_stats=trend_stats, 
                         violation_type_stats=violation_type_stats,
                         daily_resolution_stats=daily_resolution_stats,
                         history=history_log[:10])  # 초기에는 10개만 표시


@app.route('/api/generate-code', methods=['POST'])
def generate_code():
    """코드 생성 API 엔드포인트"""
    try:
        # 코드 생성 서비스 사용 가능 확인
        if not code_generator_service:
            return jsonify({
                'success': False,
                'error': 'CodeGeneratorService가 초기화되지 않았습니다. 서버 설정을 확인해주세요.'
            }), 500
        
        # 요청 데이터 검증
        data = request.json
        if not data:
            return jsonify({
                'success': False,
                'error': 'JSON 데이터가 필요합니다.'
            }), 400
        
        user_message = data.get('message', '')
        if not user_message or not user_message.strip():
            return jsonify({
                'success': False,
                'error': '메시지가 입력되지 않았습니다.'
            }), 400
        
        logger.info(f"코드 생성 요청 접수: {user_message[:100]}...")
        
        # 코드 생성 서비스 호출
        result = code_generator_service.generate_code(user_message)
        
        if result.get('success'):
            logger.info("코드 생성 성공")
            return jsonify(result)
        else:
            logger.error(f"코드 생성 실패: {result.get('error')}")
            return jsonify(result), 404 if '템플릿' in result.get('error', '') else 500
    
    except Exception as e:
        logger.error(f"코드 생성 API 처리 중 오류: {e}")
        return jsonify({
            'success': False,
            'error': f'서버 오류가 발생했습니다: {str(e)}'
        }), 500

@app.route('/api/code-generator/status', methods=['GET'])
def get_code_generator_status():
    """코드 생성 서비스 상태 확인 API"""
    try:
        if not code_generator_service:
            return jsonify({
                'status': 'error',
                'error': 'CodeGeneratorService가 초기화되지 않았습니다.'
            }), 500
        
        status = code_generator_service.get_service_status()
        return jsonify(status)
    
    except Exception as e:
        logger.error(f"서비스 상태 확인 중 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500


@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        print("최신 커밋 분석 실행 중...")
        code_validator.main()  # Assuming this function runs the validation logic
        
        # 분석 완료 후 메인 페이지로 리다이렉트
        return redirect(url_for('code_inspection'))
    except Exception as e:
        print(f"분석 중 오류 발생: {e}")
        return redirect(url_for('code_inspection'))

@app.route('/health')
def health_check():
    return 'OK', 200

@app.route('/download/excel')
def download_excel():
    """엑셀 파일 다운로드 엔드포인트"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 최근 30일간의 위반 데이터 조회
        thirty_days_ago_str = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            SELECT 
                created_at, 
                file_path, 
                author_name, 
                violation_commit_hash, 
                violation_details, 
                status, 
                resolved_at
            FROM validations 
            WHERE created_at >= ? 
            AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            ORDER BY created_at DESC
        """, (thirty_days_ago_str,))
        
        violations = cursor.fetchall()
        
        # 엑셀 파일 생성
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Validation Results"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        # 엑셀 헤더 설정
        headers = ["날짜", "파일 경로", "작성자", "커밋 해시", "위반 내용", "상태", "해결 날짜"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 추가
        for violation in violations:
            sheet.append([
                violation[0],  # created_at
                violation[1],  # file_path
                violation[2],  # author_name
                violation[3],  # violation_commit_hash
                violation[4],  # violation_details
                violation[5],  # status
                violation[6]   # resolved_at
            ])
        
        # 열 너비 자동 조정
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 파일 저장
        workbook.save(output)
        output.seek(0)
        
        # 응답 반환
        return send_file(output, 
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name="validation_results.xlsx")
    
    except Exception as e:
        logger.error(f"엑셀 다운로드 중 오류: {e}")
        return jsonify({
            'success': False,
            'error': '엑셀 파일 다운로드 중 오류가 발생했습니다.'
        }), 500

@app.route('/download-violations-excel')
def download_violations_excel():
    """위반사항 데이터를 엑셀 파일로 다운로드"""
    try:
        conn = get_db_connection()
        thirty_days_ago_str = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        
        # 모든 위반사항 데이터 조회 (가이드라인 준수 메시지 제외)
        violations_data = conn.execute("""
            SELECT 
                id,
                file_path,
                violation_commit_hash,
                author_name,
                violation_details,
                status,
                created_at,
                resolved_at,
                resolved_by_author,
                resolved_commit_hash
            FROM validations 
            WHERE violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            ORDER BY created_at DESC
        """).fetchall()
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "코드 위반사항 리포트"
        
        # 헤더 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 헤더 작성
        headers = [
            "ID", "파일 경로", "위반 커밋 해시", "작성자", "위반 내용", 
            "상태", "생성일", "해결일", "해결자", "해결 커밋 해시"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 데이터 작성
        for row_idx, violation in enumerate(violations_data, 2):
            for col_idx, value in enumerate(violation, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # 상태에 따른 색상 구분
                if col_idx == 6:  # 상태 컬럼
                    if value == 'open':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif value == 'closed':
                        cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
        
        # 컬럼 너비 자동 조정
        column_widths = [5, 40, 15, 15, 50, 10, 20, 20, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # 요약 정보 추가 (별도 시트)
        summary_ws = wb.create_sheet("요약 정보")
        
        # 통계 데이터 계산
        total_violations = len(violations_data)
        open_violations = len([v for v in violations_data if v[5] == 'open'])
        closed_violations = len([v for v in violations_data if v[5] == 'closed'])
        
        # 요약 정보 작성
        summary_data = [
            ["항목", "값"],
            ["총 위반사항", total_violations],
            ["미해결 위반사항", open_violations],
            ["해결된 위반사항", closed_violations],
            ["생성일", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # 헤더
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                cell.border = thin_border
        
        # 컬럼 너비 조정
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 15
        
        # 메모리에 엑셀 파일 저장
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        conn.close()
        
        # 파일명 생성 (현재 날짜 포함)
        filename = f"코드위반사항리포트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"엑셀 다운로드 중 오류 발생: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/activity-history')
def get_activity_history():
    """최근 활동 로그를 페이징으로 조회하는 API"""
    try:
        conn = get_db_connection()
        
        # 쿼리 파라미터 가져오기
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        offset = (page - 1) * per_page
        
        # 전체 데이터 개수 조회 (가이드라인 준수 메시지 제외)
        total_count_query = """
            SELECT COUNT(*) as total FROM (
                SELECT created_at FROM validations 
                WHERE violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
                UNION ALL
                SELECT resolved_at as created_at FROM validations 
                WHERE status = 'closed' AND resolved_at IS NOT NULL 
                AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            )
        """
        total_count = conn.execute(total_count_query).fetchone()[0]
        
        # 최근 위반사항 조회
        recent_violations = conn.execute("""
            SELECT created_at, 'open' as status, file_path, author_name, violation_commit_hash, violation_details 
            FROM validations 
            WHERE violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            ORDER BY created_at DESC 
            LIMIT ? OFFSET ?
        """, (per_page, offset)).fetchall()
        
        # 최근 해결사항 조회
        recent_resolutions = conn.execute("""
            SELECT resolved_at, 'closed' as status, file_path, resolved_by_author, resolved_commit_hash, 'Fixed' as details 
            FROM validations 
            WHERE status = 'closed' AND resolved_at IS NOT NULL 
            AND violation_details NOT LIKE '%코드 가이드라인을 모두 준수했습니다%'
            ORDER BY resolved_at DESC 
            LIMIT ? OFFSET ?
        """, (per_page, offset)).fetchall()
        
        # 두 리스트를 합치고 시간 순으로 정렬
        history_log = []
        
        for row in recent_violations:
            if row[0]:  # created_at이 None이 아닌 경우
                try:
                    timestamp = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
                    history_log.append({
                        'timestamp': timestamp, 
                        'status': row[1], 
                        'file_path': row[2], 
                        'author': row[3], 
                        'commit_hash': row[4], 
                        'details': row[5]
                    })
                except ValueError:
                    continue
                    
        for row in recent_resolutions:
            if row[0]:  # resolved_at이 None이 아닌 경우
                try:
                    timestamp = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
                    history_log.append({
                        'timestamp': timestamp, 
                        'status': row[1], 
                        'file_path': row[2], 
                        'author': row[3], 
                        'commit_hash': row[4], 
                        'details': row[5]
                    })
                except ValueError:
                    continue
        
        # 시간 순으로 정렬하고 페이징 적용
        history_log.sort(key=lambda x: x['timestamp'], reverse=True)
        paginated_history = history_log[:per_page]
        
        # 응답 데이터 구성
        response_data = {
            'success': True,
            'data': [{
                'timestamp': item['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                'status': item['status'],
                'file_path': item['file_path'],
                'author': item['author'],
                'commit_hash': item['commit_hash'],
                'details': item['details']
            } for item in paginated_history],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total_count,
                'has_next': (page * per_page) < total_count,
                'total_pages': (total_count + per_page - 1) // per_page
            }
        }
        
        conn.close()
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"활동 히스토리 조회 중 오류 발생: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8000)
